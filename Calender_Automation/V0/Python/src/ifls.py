"""To Generate Integrated Faculty Load Sheet"""
import os
from openpyxl import load_workbook
from faculty_class import InputSchedule, FacultySlots

# Function to Update Faculty Session Slots
def update_faculty(name, session,faculty_list):
    """For each faculty, update the count of slots"""
    check = 'No Name'
    if name is not None:
        for i in range(len(faculty_list)):
            if name == faculty_list[i].get_faculty_name():
                if session == 'M':
                    faculty_list[i].update_m_slots()
                    check = 'Morning'
                elif session == 'A':
                    faculty_list[i].update_a_slots()
                    check = 'Afternoon'
                elif session == 'M&A':
                    faculty_list[i].update_m_slots()
                    faculty_list[i].update_a_slots()
                    check = 'Morning and Afternoon'
                else:
                    check = 'No Session'
        return check
    return check

#Function to Check Avaialble slots and return
def free_slots(slots):
    """With planned slots as 18 return available slots"""
    return 18 - slots

#Function to generate the output sheet
def create_ifls(path, inp):
    """Function to create IFLS sheet"""
    work_book = load_workbook(path)
    work_sheet = work_book.active
    input_list = []
    faculty_list = []

    program_initiative = work_sheet['A1'].value
    schedule_month = work_sheet['A4'].value

    for cols in range(0, 31):
        lead_1 = work_sheet['F' + str(cols + 4)].value
        lead_2 = work_sheet['G' + str(cols + 4)].value
        lead_3 = work_sheet['H' + str(cols + 4)].value
        session_slot = work_sheet['I' + str(cols + 4)].value
        temp = InputSchedule(lead_1, lead_2, lead_3, session_slot)
        input_list.append(temp)

    temp_list = []
    def check_unique_value(fac_name):
        """Add unique faculty names onto the list"""
        if fac_name is not None and not fac_name in temp_list:
            faculty = FacultySlots(fac_name)
            temp_list.append(fac_name)
            faculty_list.append(faculty)

    for i in range(len(input_list)):
        l_1 = input_list[i].get_lead_1()
        l_2 = input_list[i].get_lead_2()
        l_3 = input_list[i].get_lead_3()
        check_unique_value(l_1)
        check_unique_value(l_2)
        check_unique_value(l_3)

    for i in range(len(input_list)):
        fac_name_1 = input_list[i].get_lead_1()
        fac_name_2 = input_list[i].get_lead_2()
        fac_name_3 = input_list[i].get_lead_3()
        session = input_list[i].get_session_slot()
        temp_var = update_faculty(fac_name_1, session, faculty_list)
        if fac_name_2 != fac_name_1:
            temp_var = update_faculty(fac_name_2, session, faculty_list)
        if fac_name_3 not in (fac_name_1, fac_name_2):
            temp_var = update_faculty(fac_name_3, session, faculty_list)

    # Output Sheet Creation
    ifls = inp.create_sheet('Integrated Faculty Load Sheet')

    ifls.merge_cells('C1:H1')
    ifls['C1'] = schedule_month + ' 2021 Faculty Load '

    ifls.merge_cells('C2:D2')
    ifls['C2'] = program_initiative

    ifls.merge_cells('E2:F2')
    ifls['E2'] = 'Planned Slots'

    ifls.merge_cells('G2:H2')
    ifls['G2'] = 'Available Slots'

    headings = ['S.No','Faculty Name', 'M', 'A', 'M', 'A', 'M', 'A']
    ifls.append(headings)

    for i in range(len(faculty_list)):
        name = faculty_list[i].get_faculty_name()
        mslots = faculty_list[i].get_m_slots()
        aslots = faculty_list[i].get_a_slots()
        row = [i+1, name, mslots, aslots, 18, 18, free_slots(mslots), free_slots(aslots)]
        ifls.append(row)

    directory = 'Result Calendar'
    file = 'Faculty Calendar.xlsx'
    if not os.path.exists(directory):
        os.makedirs(directory)

    inp.save(os.path.join(directory, file))
    return 1

