'''
Genrate Faculty Calendar sheet
importing required modules
Pylint score = 9.50/10
'''
import os
from openpyxl.styles.alignment import Alignment
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from faculty_class import InputSchedule, FacultySlots

def create_faculty_calendar(path,inp):
    """
    Genrating Faculty Calendar
    """
    inputfile = load_workbook(path)
    inputsheet = inputfile.active
    #new worksheet
    faculty_sheet = inp.active
    faculty_sheet.title = 'Faculty sheet'
    inputlist_fl = []
    facultylist_fl = []
    date=[i for i in range(1,32)]
    schedule_month = inputsheet['A4'].value
    for coloums in range(0, 31):
        lead1 = inputsheet['F' + str(coloums + 4)].value
        lead2 = inputsheet['G' + str(coloums + 4)].value
        lead3 = inputsheet['H' + str(coloums + 4)].value
        sesessionslotion_slot = inputsheet['I' + str(coloums + 4)].value
        temp = InputSchedule(lead1, lead2, lead3, sesessionslotion_slot)
        inputlist_fl.append(temp)

    temp_list = []
    def finduniquevalue(fac_name):
        if fac_name is not None and not fac_name in temp_list:
            faculty = FacultySlots(fac_name)
            temp_list.append(fac_name)
            facultylist_fl.append(faculty)

    for i in inputlist_fl:
        temp1 = i.get_lead_1()
        temp2 = i.get_lead_2()
        temp3 = i.get_lead_3()
        finduniquevalue(temp1)
        finduniquevalue(temp2)
        finduniquevalue(temp3)
    namestore=5      #for faculty name storage
    rowiter=5        #for storring values "1" or "0" in respective cell

    #Output display format of sheet
    heading_merge(faculty_sheet,schedule_month)
    date_align(faculty_sheet,date)
    session_align(faculty_sheet)
    #Computation of Faculty Calendar
    for i in facultylist_fl:
        name = i.get_faculty_name()
        faculty_sheet['A'+str(namestore)]=name
        namestore+=1
    for i in facultylist_fl:
        coloum=2
        name= i.get_faculty_name()
        for j in inputlist_fl:
            given1 =j.get_lead_1()
            given2 =j.get_lead_2()
            given3 =j.get_lead_3()
            sessionslot =j.get_session_slot()

            given2=check(given2)
            given3=check(given3)
            length_sessionslot=lengthofsession(sessionslot)
            iteration=2
            while iteration:
                char=get_column_letter(coloum)
                if length_sessionslot==0:
                    faculty_sheet[char +str(rowiter)]='Holiday'
                if length_sessionslot>1:
                    if((name is given1) or (name is given2) or (name is given3)):
                        faculty_sheet[char + str(rowiter)]='1'
                    else:
                        faculty_sheet[char + str(rowiter)]='0'
                else:
                    if sessionslot=='M':
                        if(((name is given1)or(name is given2)or(name is given3))and(coloum%2==0)):
                            faculty_sheet[char + str(rowiter)]='1'
                        else:
                            faculty_sheet[char + str(rowiter)]='0'
                    if sessionslot=='A':
                        if(((name is given1)or(name is given2)or(name is given3))and(coloum%2!=0)):
                            faculty_sheet[char + str(rowiter)]='1'
                        else:
                            faculty_sheet[char + str(rowiter)]='0'
                coloum+=1
                iteration-=1
        rowiter+=1
    # Saving the file
    directory = 'Result Calendar'
    file = 'Faculty Calendar.xlsx'
    if not os.path.exists(directory):
        os.makedirs(directory)

    inp.save(os.path.join(directory, file))
    return 1


def lengthofsession(sesessionslotion_solt):
    '''
    Calculate lenght of sesessionslotion slot
    '''
    if sesessionslotion_solt is not None:
        return len(sesessionslotion_solt)
    return 0
def check(given):
    '''
    Check if input is None
    '''
    if given is None:
        return "empty"
    return given
def date_align(faculty_sheet,date):
    '''
    Align date in order to store in sheet
    '''
    i=0
    for coloum in range(2,64,2):
        char=get_column_letter(coloum)
        faculty_sheet[char +'2']=str(date[i])
        i+=1
def session_align(faculty_sheet):
    '''
    Align sesessionslotion cell store A/m in respective cell
    '''
    for coloum in range(2,64,1):
        char=get_column_letter(coloum)
        if coloum%2==0:
            faculty_sheet[char +'3']='M'
        else:
            faculty_sheet[char +'3']='A'
def heading_merge(faculty_sheet, schedule_month):
    '''
    Heading required to display in sheet and merge cells
    '''

    faculty_sheet.merge_cells('B1:BK1')
    faculty_sheet['B1'] = schedule_month + '2021 Faculty Calendar'
    faculty_sheet['B1'].alignment=Alignment(horizontal='center')
    faculty_sheet.column_dimensions['A'].width=30
    faculty_sheet['A1']='Month'
    faculty_sheet['A2']='Date'
    faculty_sheet['A3']= 'Sesessionslotion [M- morning, A- Afternoon]'
    faculty_sheet['A4']='Faculty Name'

    faculty_sheet.merge_cells('B2:C2')
    faculty_sheet.merge_cells('D2:E2')
    faculty_sheet.merge_cells('F2:G2')
    faculty_sheet.merge_cells('H2:I2')
    faculty_sheet.merge_cells('J2:K2')
    faculty_sheet.merge_cells('L2:M2')
    faculty_sheet.merge_cells('N2:O2')
    faculty_sheet.merge_cells('P2:Q2')
    faculty_sheet.merge_cells('R2:S2')
    faculty_sheet.merge_cells('T2:U2')
    faculty_sheet.merge_cells('V2:W2')
    faculty_sheet.merge_cells('X2:Y2')
    faculty_sheet.merge_cells('Z2:AA2')
    faculty_sheet.merge_cells('AB2:AC2')
    faculty_sheet.merge_cells('AD2:AE2')
    faculty_sheet.merge_cells('AF2:AG2')
    faculty_sheet.merge_cells('AH2:AI2')
    faculty_sheet.merge_cells('AJ2:AK2')
    faculty_sheet.merge_cells('AL2:AM2')
    faculty_sheet.merge_cells('AN2:AO2')
    faculty_sheet.merge_cells('AP2:AQ2')
    faculty_sheet.merge_cells('AR2:AS2')
    faculty_sheet.merge_cells('AT2:AU2')
    faculty_sheet.merge_cells('AV2:AW2')
    faculty_sheet.merge_cells('AX2:AY2')
    faculty_sheet.merge_cells('AZ2:BA2')
    faculty_sheet.merge_cells('BB2:BC2')
    faculty_sheet.merge_cells('BD2:BE2')
    faculty_sheet.merge_cells('BF2:BG2')
    faculty_sheet.merge_cells('BH2:BI2')
    faculty_sheet.merge_cells('BJ2:BK2') #end
