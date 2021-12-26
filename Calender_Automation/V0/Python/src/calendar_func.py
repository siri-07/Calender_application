"""
Import required packages namely pandas and openpyxl
"""
import os
from string import ascii_uppercase
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side


class MasterCalendar:
    """Class for master calendar"""

    def __init__(self):
        self.flag = 0

    def create_master_cal(self, path):
        """function to create master calendar"""
        header_df = pd.read_excel(path, sheet_name='Sheet1')
        training = header_df.columns[0]
        track = header_df[0:1]
        track_name = (track[training][0])
        header_df.columns = ['Month', 'Date', 'Day', 'Course Code', 'Module',
                             'Lead1', 'Lead2', 'Lead3', 'Session Slot', 'Session Time']
        header_df = header_df.drop([0, 1])
        header_df.dropna(subset=["Course Code"], inplace=True)
        header_df = header_df.drop(['Month', 'Date', 'Day', 'Session Slot', 'Session Time'], axis=1)
        result = header_df.drop_duplicates()
        session_df = pd.read_excel(path, sheet_name='Sheet1')
        session_df.columns = ['Month', 'Date', 'Day', 'Course Code',
                              'Module', 'Lead1', 'Lead2', 'Lead3',
                              'Session Slot', 'Session Time']
        session_df = session_df.drop([0, 1])
        session_df = session_df.drop(['Module', 'Lead1',
                                      'Lead2', 'Lead3', 'Session Time'], axis=1)
        session_df.set_index('Date')
        courses = result['Course Code'].tolist()
        month = set(session_df['Month'])
        days = session_df['Day']
        name_of_month = month.pop()
        name1, name2 = "M", "A"
        date = 2
        for i in range(5, (len(days) * 2) + 5):
            rowval = [0, 0, 0, 0, 0]
            if i % 2 != 0:
                if session_df['Session Slot'][date] == 'M' \
                        or session_df['Session Slot'][date] == 'M&A':
                    for j in range(len(courses)):
                        if courses[j] == session_df['Course Code'][date]:
                            rowval[j] = 1
            if i % 2 == 0:
                if session_df['Session Slot'][date - 1] == 'A' \
                        or session_df['Session Slot'][date - 1] == 'M&A':
                    for j in range(len(courses)):
                        if courses[j] == session_df['Course Code'][date - 1]:
                            rowval[j] = 1
            if i % 2 == 0:
                result.insert(i, name2, rowval, allow_duplicates=True)
            else:
                date = date + 1
                result.insert(i, name1, rowval, allow_duplicates=True)
        directory = 'Result Calendar'
        file = 'Master_Calendar.xlsx'
        if not os.path.exists(directory):
            os.makedirs(directory)
        result.to_excel(os.path.join(directory, file), index=False, startrow=2)
        output_excel = load_workbook(os.path.join(directory, file))
        master_calender_sheet = output_excel["Sheet1"]
        master_calender_sheet.merge_cells('F1:BO1')
        master_calender_sheet['F1'] = name_of_month
        master_calender_sheet['A1'] = training
        master_calender_sheet.merge_cells('A1:B1')
        master_calender_sheet['A1'].fill = PatternFill(fill_type='solid', start_color='0000FFFF',
                                                       end_color='0000FFFF')
        master_calender_sheet['A2'] = track_name
        master_calender_sheet.merge_cells('A2:B2')
        master_calender_sheet['A2'].fill = PatternFill(fill_type='solid', start_color='00808000',
                                                       end_color='00808000')
        cell = master_calender_sheet['F1']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        set_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        for row in master_calender_sheet.iter_rows():
            for cell in row:
                if cell.value == 1:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='00FF0000', end_color='00FF0000')
                    cell.border = set_border
                    cell.value = ''
                if cell.value == 0:
                    cell.value = ''
                if cell.value == name_of_month:
                    cell.fill = PatternFill(fill_type='solid',
                                            start_color='000000FF', end_color='000000FF')
        day = 1
        check_same_day = 0
        for row in master_calender_sheet.iter_cols(min_row=2, max_row=2, min_col=6, max_col=67):
            # for each row there is one cell object (since min_col = max_col)
            for cell in row:
                if check_same_day == 0 or check_same_day == 1:
                    cell.value = day
                    check_same_day += 1
                else:
                    cell.value = day + 1
                    day += 1
                    check_same_day = 1

        lst = list(ascii_uppercase) + \
              [letter1 + letter2 for letter1 in ascii_uppercase for letter2 in ascii_uppercase]
        arr = lst[5:]
        for i in range(len(days) * 2):
            if i % 2 == 0:
                master_calender_sheet.merge_cells(str(arr[i]) + '2:' + str(arr[i + 1]) + '2')
                cell = master_calender_sheet[str(arr[i + 1]) + '2']
                cell.border = set_border
                cell = master_calender_sheet[str(arr[i]) + '2']
                cell.border = set_border
                cell.fill = PatternFill(fill_type='solid', start_color='00FFFF00',
                                        end_color='00FFFF00')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(fill_type='solid', start_color='00FFFF00',
                                        end_color='00FFFF00')
                cell.border = set_border
        output_excel.save(os.path.join(directory, file))
        self.flag = 1
        return self.flag
