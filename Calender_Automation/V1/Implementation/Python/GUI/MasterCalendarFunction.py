# importing the dependencies
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side
from pandas.core.frame import DataFrame
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import datetime

"""
Template for google drive operations like upload, download, search for files
"""
gauth = GoogleAuth()
gauth.LoadCredentialsFile("credentials.txt")
drive = GoogleDrive(gauth)


def upload_master_calendar(file_name):
    """Upload file -> change name in setcontentfile"""
    file1 = drive.CreateFile({"mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    file1.SetContentFile(file_name)
    file1.Upload({"convert": True})

    folder_name = 'Download'

    folders = drive.ListFile(
        {'q': "title='" + folder_name + "' and mimeType='application/vnd.google-apps.folder' and trashed=false"}).GetList()
    for folder in folders:
        if folder['title'] == folder_name:
            file2 = drive.CreateFile({'parents': [{'id': folder['id']}]})
            file2.SetContentFile(file_name)
            file2.Upload()



def download_master_calendar():
    """
    Searching for a file and downloading it -> insert file name after contains
    give file name to variable file_name
    """
    file_list = drive.ListFile({'q': "title contains 'Master' and trashed=false"}).GetList()
    # print(file_list[0]['title'])
    file_id = file_list[0]['id']
    # print(file_id)
    file = drive.CreateFile({'id': file_id})
    file_time = datetime.now().strftime(" %Y-%m-%d_%I-%M-%S_%p")
    file_name = 'Master'
    file_format = '.xlsx'
    file_title = file_name + file_time + file_format
    file.GetContentFile(file_title,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return file_title


def MasterCalendarFunction(SheetName, InputPath, Initiative, OutMonth):
    """Extracting data from the input calender
    which is in the day wise format
    """
    InputDataframe = pd.read_excel(InputPath, sheet_name=SheetName)
    InputDataframe.columns = ['Month', 'Date', 'Day', 'Course Code', 'Module', 'Lead1', 'Lead2', 'Lead3',
                              'Session Slot',
                              'Session Time', 'Comments']
    InputDataframe = InputDataframe.drop([0, 1])
    InputDataframe.reset_index(inplace=True, drop=True)
    Date = InputDataframe['Date']
    CourseCode = InputDataframe['Course Code']
    Module = InputDataframe['Module']
    Lead1 = InputDataframe['Lead1']
    Lead2 = InputDataframe['Lead2']
    Lead3 = InputDataframe['Lead3']
    SessionSlot = InputDataframe['Session Slot']

    """Extracting data from the Keys
    sheet of the Master calendar"""
    ExistingMaster = download_master_calendar()
    KeysDataframe = pd.read_excel(InputPath, sheet_name='Key')
    KeysDataframe.columns = ["FixedInitiativeTitles", "FixedInitiativeCodes", "VarName3", "VarName4",
                             "VarName5", "VarName6", "FixedCourseCodes", "FixedCourseTitles"]
    KeysDataframe = KeysDataframe.drop(["VarName4", "VarName5", "VarName6"], axis=1)
    FixedInitiativeTitles = KeysDataframe['FixedInitiativeTitles']
    FixedInitiativeCodes = KeysDataframe['FixedInitiativeCodes']
    FixedCourseCodes = KeysDataframe['FixedCourseCodes']
    FixedCourseTitles = KeysDataframe['FixedCourseTitles']

    """Error correction :
    if course code incorrect , course title correct         =   corrects course code
    if course code correct   , course title  incorrect      =   corrects course title
    if course code incorrect , course title also incorrect  =   replaces the course code with ""
    """

    """Removing all the leading and trailing spaces from CourseCode, Module, FixedInitiativeTitles, FixedCourseCodes, 
    fixedCourseTitles """
    CourseCode.str.strip()
    Module.str.strip()
    FixedInitiativeTitles.str.strip()
    FixedCourseCodes.str.strip()
    FixedCourseTitles.str.strip()

    """Fixing the error course codes"""
    for i in range(0, len(CourseCode)):
        TempFlag = 0
        for j in range(0, len(FixedCourseCodes)):
            if CourseCode[i] == FixedCourseCodes[j]:
                TempFlag = 1
        if TempFlag == 0:
            TempFlagError = 1
            for k in range(0, len(FixedCourseTitles)):
                if Module[i] == FixedCourseTitles[k]:
                    CourseCode[i] = FixedCourseCodes[k]
                    TempFlagError = 0
            if TempFlagError == 1:
                CourseCode[i] = ""

    "Fixing the error course titles"
    for i in range(0, len(Module)):
        TempFlag = 0
        for j in range(0, len(FixedCourseTitles)):
            if Module[i] == FixedCourseTitles[j]:
                TempFlag = 1
        if TempFlag == 0:
            TempFlagError = 1
            for k in range(0, len(FixedCourseCodes)):
                if CourseCode[i] == FixedCourseCodes[k]:
                    Module[i] = FixedCourseTitles[k]
                    TempFlagError = 0
            if TempFlagError == 1:
                CourseCode[i] = ""

    """Selecting the particular initaitive code"""
    InitiativeCode = 11
    for i in range(0, len(FixedInitiativeTitles)):
        if Initiative == FixedInitiativeTitles[i]:
            InitiativeCode = FixedInitiativeCodes[i]

    """UniqueCourseCode containing unique data for CourseCode"""
    UniqueCourseCode = []
    for i in range(0, len(CourseCode)):
        if CourseCode[i] != '' and CourseCode[i] not in UniqueCourseCode:
            UniqueCourseCode.append(CourseCode[i])
    UniqueCourseCode = pd.Series(UniqueCourseCode)

    """Data containing the date-wise module names and respective faculties"""
    Data = pd.concat([Module, Lead1, Lead2, Lead3], axis=1)  # Module-Lead1-Lead2-Lead3 dataframe

    """Declaring variable to hold respective CourseTitle for UniqueCourseCode"""
    RespectiveCourseTitle = pd.Series([""] * len(UniqueCourseCode))

    """Declaring matrix to hold repeatitive list of faculties for respective UniqueCourseCode"""
    Faculty = pd.DataFrame([[""] * len(CourseCode) * 3] * len(UniqueCourseCode))

    """Declaring matrix to hold respective list of
    faculties for respective UniqueCourseCode"""
    RespectiveFaculty = pd.DataFrame([[""] * 5] * len(UniqueCourseCode))

    """Initialising a TimeTable of zeros for UniqueCourseCode for a month of 31 days
    """
    TimeTable = pd.DataFrame([[0] * 124] * len(UniqueCourseCode))

    """Logically assigning a CourseTitle,
    Faculty for every UniqueCourseCode"""
    for i in range(len(UniqueCourseCode)):
        for j in range(len(CourseCode)):
            if UniqueCourseCode[i] == CourseCode[j]:
                RespectiveCourseTitle[i] = Data.iloc[j, 0]
                Faculty.iloc[i, (j * 3):(j * 3) + 3] = Data.iloc[j, 1:4]
                if SessionSlot[j] == 'M' or SessionSlot[j] == 'm':
                    TimeTable.iloc[i, (4 * Date[j] - 4)] = InitiativeCode
                    TimeTable.iloc[i, (4 * Date[j] - 3)] = InitiativeCode
                elif SessionSlot[j] == 'A' or SessionSlot[j] == 'a':
                    TimeTable.iloc[i, (4 * Date[j] - 2)] = InitiativeCode
                    TimeTable.iloc[i, (4 * Date[j] - 1)] = InitiativeCode
                elif SessionSlot[j] == 'F' or SessionSlot[j] == 'f':
                    TimeTable.iloc[i, (4 * Date[j] - 4)] = InitiativeCode
                    TimeTable.iloc[i, (4 * Date[j] - 3)] = InitiativeCode
                    TimeTable.iloc[i, (4 * Date[j] - 2)] = InitiativeCode
                    TimeTable.iloc[i, (4 * Date[j] - 1)] = InitiativeCode
                elif SessionSlot[j] == 'M1' or SessionSlot[j] == 'm1':
                    TimeTable.iloc[i, (4 * Date[j] - 4)] = InitiativeCode
                elif SessionSlot[j] == 'M2' or SessionSlot[j] == 'm2':
                    TimeTable.iloc[i, (4 * Date[j] - 3)] = InitiativeCode
                elif SessionSlot[j] == 'A1' or SessionSlot[j] == 'a1':
                    TimeTable.iloc[i, (4 * Date[j] - 2)] = InitiativeCode
                elif SessionSlot[j] == 'A2' or SessionSlot[j] == 'a2':
                    TimeTable.iloc[i, (4 * Date[j] - 1)] = InitiativeCode

    """Replacing all the NaN in Faculty with "" empty strings"""
    NanValue = float("NaN")
    Faculty.replace(NanValue, "", inplace=True)

    """Converting everything in Faculty to upper case"""
    Faculty = Faculty.apply(lambda x: x.astype(str).str.upper())

    """removing leading and trailing spaces from everything in Faculty"""
    Faculty = Faculty.apply(lambda x: x.astype(str).str.strip())

    """Replace all the "" in Faculty with NaN"""
    NanValue = float("NaN")
    Faculty.replace("", NanValue, inplace=True)

    """Forming the list of unique faculties for a particular course and saving it in RespectiveFaculty"""
    for i in range(0, len(UniqueCourseCode)):
        UniqueFaculty = pd.Series(Faculty.iloc[i, :].unique())
        UniqueFaculty.dropna(inplace=True)
        RespectiveFaculty.iloc[i, 0:len(UniqueFaculty)] = UniqueFaculty

    """
    Importing the existing data from OutMonth.xlsx sheet of Master.xlsx workbook
    """
    ExistingDataframe = pd.read_excel(ExistingMaster, sheet_name=OutMonth)
    ExistingDataframe = DataFrame(ExistingDataframe.drop([0, 1]))
    ExistingDataframe = ExistingDataframe.reset_index(drop=True)
    UniqueCourseCodeOutMonth = pd.Series(ExistingDataframe.iloc[:, 0])
    RespectiveCourseTitleOutMonth = pd.Series(ExistingDataframe.iloc[:, 1])
    RespectiveFacultyOutMonth = DataFrame(ExistingDataframe.iloc[:, 2:6])
    TimeTableOutMonth = DataFrame(ExistingDataframe.iloc[:, 7:131])
    RespectiveFacultyWidthOutMonth = len(RespectiveFacultyOutMonth.columns)

    """Calculating total number of days in a month"""
    if ((OutMonth == 'January') or (
            OutMonth == 'March') or OutMonth == 'May' or OutMonth == 'July' or OutMonth == 'August' or OutMonth == 'October' or OutMonth == 'December'):
        TotalDays = 124
    elif OutMonth == 'February':
        TotalDays = 116
    else:
        TotalDays = 120

    """
    Opening Master calendar excel using openpyxl
    """
    WriteExcel = load_workbook(ExistingMaster)
    MasterCalendarOutMonth = WriteExcel[OutMonth]
    """
    Checking the data that already exists on OutMonth.xlsx sheet of Master.xlsx workbook
    If data == Empty  , the over-write the new data (OutputDataFrame)
    If data != Empty  , then append the new data (OutputDataFrame) to the existing data
    """
    IsEmpty = UniqueCourseCodeOutMonth.empty
    OutputDataframe = pd.concat([UniqueCourseCode, RespectiveCourseTitle, RespectiveFaculty], axis=1, ignore_index=True)
    FinalLength = len(UniqueCourseCode)
    if IsEmpty:
        rows = dataframe_to_rows(OutputDataframe, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 1):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
        Schedule = dataframe_to_rows(TimeTable, index=False, header=False)
        for r_idx, row in enumerate(Schedule, 4):
            for c_idx, value in enumerate(row, 8):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
    else:
        """Counting the number of common courses in between UniqueCourseCode and UniqueCourseCodeOutMonth"""
        TempCounterFinal = 0
        for i in range(0, len(UniqueCourseCode)):
            for j in range(0, len(UniqueCourseCodeOutMonth)):
                if UniqueCourseCode[i] == UniqueCourseCodeOutMonth[j]:
                    TempCounterFinal = TempCounterFinal + 1

        """Initialising the final outputs of FinalUniqueCourseCode, FinalRespectiveCourseTitle, 
        FinalRespectiveFaculty """
        FinalLength = len(UniqueCourseCodeOutMonth) + len(UniqueCourseCode) - TempCounterFinal

        FinalUniqueCourseCode = pd.Series([""] * FinalLength)
        FinalUniqueCourseCode[:len(UniqueCourseCodeOutMonth)] = UniqueCourseCodeOutMonth

        FinalRespectiveCourseTitle = pd.Series([""] * FinalLength)
        FinalRespectiveCourseTitle[:len(UniqueCourseCodeOutMonth)] = RespectiveCourseTitleOutMonth

        FinalRespectiveFaculty = pd.DataFrame([[""] * 5] * FinalLength)
        FinalRespectiveFaculty.iloc[:len(UniqueCourseCodeOutMonth),
        0:RespectiveFacultyWidthOutMonth] = RespectiveFacultyOutMonth
        FinalRespectiveFaculty.replace(NanValue, "", inplace=True)

        FinalTimeTable = pd.DataFrame([[0] * TotalDays] * FinalLength)
        FinalTimeTable.iloc[:len(UniqueCourseCodeOutMonth), :] = TimeTableOutMonth
        TempCounterCourse = 0
        TempCounterFaculty = 0
        for i in range(0, len(UniqueCourseCode)):
            TempFlagCourse = 0
            for j in range(0, len(FinalUniqueCourseCode)):
                if UniqueCourseCode[i] == FinalUniqueCourseCode[j]:
                    TempFlagCourse = 1
                    TempRow = j
            if TempFlagCourse == 1:
                RespectiveFacultyLength = pd.Series(RespectiveFaculty.iloc[i, :].unique())
                RespectiveFacultyLength.replace("", NanValue, inplace=True)
                RespectiveFacultyLength.dropna(inplace=True)
                RespectiveFacultyLength = len(RespectiveFacultyLength)
                FinalRespectiveFacultyLength = pd.Series(FinalRespectiveFaculty.iloc[TempRow, :].unique())
                FinalRespectiveFacultyLength.replace("", NanValue, inplace=True)
                FinalRespectiveFacultyLength.dropna(inplace=True)
                FinalRespectiveFacultyLength = len(FinalRespectiveFacultyLength)
                for x in range(0, RespectiveFacultyLength):
                    TempFlagFaculty = 0
                    for y in range(0, FinalRespectiveFacultyLength):
                        if RespectiveFaculty.iloc[i, x] == FinalRespectiveFaculty.iloc[TempRow, y]:
                            TempFlagFaculty = 1
                    if TempFlagFaculty == 0:
                        FinalRespectiveFaculty.iloc[TempRow, FinalRespectiveFacultyLength + TempCounterFaculty] = \
                            RespectiveFaculty.iloc[i, x]
                        TempCounterFaculty = TempCounterFaculty + 1
                    if (TempCounterFaculty + FinalRespectiveFacultyLength) > 4:
                        TempCounterFaculty = 0
                        ErrorCode = 1
                        break
                for a in range(0, 123):
                    if TimeTable.iloc[i, a] == InitiativeCode:
                        FinalTimeTable.iloc[TempRow, a] = TimeTable.iloc[i, a]
            elif TempFlagCourse == 0:
                TempRowFinal = len(UniqueCourseCodeOutMonth) + TempCounterCourse
                FinalUniqueCourseCode[TempRowFinal] = UniqueCourseCode[i]
                FinalRespectiveCourseTitle[TempRowFinal] = RespectiveCourseTitle[i]
                FinalRespectiveFaculty.iloc[TempRowFinal, :] = RespectiveFaculty.iloc[i, :]
                FinalTimeTable.iloc[TempRowFinal, :] = TimeTable.iloc[i, :]
                TempCounterCourse = TempCounterCourse + 1
                if TempRowFinal > (len(FixedCourseCodes) - 1):
                    ErrorCode = 2
                    break
        FinalDataframe = pd.concat([FinalUniqueCourseCode, FinalRespectiveCourseTitle, FinalRespectiveFaculty], axis=1,
                                   ignore_index=True)
        rows = dataframe_to_rows(FinalDataframe, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 1):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
        Schedule = dataframe_to_rows(FinalTimeTable, index=False, header=False)
        for r_idx, row in enumerate(Schedule, 4):
            for c_idx, value in enumerate(row, 8):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)

    """
    Setting colour codes for Initiatives 
    """
    KeysDataframe = KeysDataframe.drop(["FixedCourseCodes", "FixedCourseTitles"], axis=1)
    KeySheet = WriteExcel["Key"]
    KeySheet.delete_rows(1, KeySheet.max_row)
    KeySheet.delete_cols(1, KeySheet.max_column)
    rows = dataframe_to_rows(KeysDataframe, index=False, header=True)
    for r_idx, row in enumerate(rows,1):
        for c_idx, value in enumerate(row,1):
            KeySheet.cell(row=r_idx, column=c_idx, value=value)

    KeyCodes = list(FixedInitiativeTitles)
    ColourValues = ['009EE362', '0000C0D0', '00FFD403', '00FF9356', '007E74D4', '00FE82AA', '00B28DFF', '0085E3FF',
                    '00BFFCC6', '00E7FFAC', '00B5D8D6', '00F6E7E0', '0033CCCC', '00FFCC99', '00333399', '00CC99FF',
                    '00FF00FF', '00FFFF00', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF']

    """
    Fill the colour for cells in colour code column
    """
    i = 0
    FixedInitiativeCodes = FixedInitiativeCodes.dropna()
    heading = KeySheet.cell(row=1,column=3)
    heading.value = "Initiative Color Code"

    for row in KeySheet.iter_rows(min_row=2, min_col=3, max_row=len(FixedInitiativeCodes) + 1, max_col=3):
        for cell in row:
            cell.fill = PatternFill(fill_type='solid',
                                    start_color=ColourValues[i], end_color=ColourValues[i])
            i = i + 1

    KeyDictionary = {}
    for i in range(len(KeyCodes)):
        KeyDictionary[KeyCodes[i]] = ColourValues[i]
    set_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    """
     Iterate through timetable cell values and fill if value is course code else blank
    """
    FixedInitiativeTitles = KeysDataframe['FixedInitiativeTitles']
    FixedInitiativeTitles = FixedInitiativeTitles.dropna()
    FixedInitiativeCodes = KeysDataframe['FixedInitiativeCodes']
    FixedInitiativeCodes = FixedInitiativeCodes.dropna()
    InitiativeKey = {}
    for i in range(len(FixedInitiativeTitles)):
        InitiativeKey[FixedInitiativeTitles[i]] = (FixedInitiativeCodes[i])
    """Fill colour on session slots"""
    for row in MasterCalendarOutMonth.iter_rows(min_row=4, min_col=8):
        for cell in row:
            if cell.value == InitiativeKey[Initiative]:
                cell.fill = PatternFill(fill_type='solid',
                                        start_color=KeyDictionary[Initiative], end_color=KeyDictionary[Initiative])
                cell.border = set_border
            if cell.value == 0:
                cell.value = ''
    """Dates not analysed"""
    DatesNotAnalysed = []
    for i in range(len(CourseCode)):
        if CourseCode[i] == "":
            DatesNotAnalysed.append(Date[i])
    print(DatesNotAnalysed)

    """ Setting the borders for each cell"""
    for i in range(1, FinalLength + 4):
        for j in range(1, TotalDays + 8):
            cell_obj = MasterCalendarOutMonth.cell(row=i, column=j)
            cell_obj.border = set_border

    """Saving the updated Master.xlsx workbook"""
    WriteExcel.save(ExistingMaster)
    upload_master_calendar(ExistingMaster)