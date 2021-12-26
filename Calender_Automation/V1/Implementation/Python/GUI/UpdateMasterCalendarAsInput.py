# importing the dependencies
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side
from pandas.core.frame import DataFrame
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import datetime

"""
Template for google drive operations like upload, download, search for files
"""
gauth = GoogleAuth()
gauth.LoadCredentialsFile("credentials.txt")
drive = GoogleDrive(gauth)

"""
Function to upload to drive the master calendar
"""


def upload_master_calendar(file_name):
    """Upload file -> change name in setcontentfile"""
    file1 = drive.CreateFile({"mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    file1.SetContentFile(file_name)
    file1.Upload({"convert": True})

    folder_name = 'Download'

    folders = drive.ListFile(
        {
            'q': "title='" + folder_name + "' and mimeType='application/vnd.google-apps.folder' and trashed=false"}).GetList()
    for folder in folders:
        if folder['title'] == folder_name:
            file2 = drive.CreateFile({'parents': [{'id': folder['id']}]})
            file2.SetContentFile(file_name)
            file2.Upload()


"""
Function to download the latest copy of master calendar from the file and update it
"""


def download_master_calendar():
    """
    Searching for a file and downloading it -> insert file name after contains
    give file name to variable file_name
    """
    file_list = drive.ListFile({'q': "title contains 'Master' and trashed=false"}).GetList()
    file_id = file_list[0]['id']
    file = drive.CreateFile({'id': file_id})
    file_time = datetime.now().strftime(" %Y-%m-%d_%I-%M-%S_%p")
    file_name = 'Master'
    file_format = '.xlsx'
    file_title = file_name + file_time + file_format
    file.GetContentFile(file_title,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return file_title


"""
Function to update the master calendar when the input file is Master calendar 
Input calendar maybe of form two slots - M/A or four slots - M1,M2,A1,A2
"""


def UpdateMasterCalendarAsInput(input_path, month):
    # read input file master calendar
    input_dataframe = pd.read_excel(input_path, sheet_name=month)
    input_dataframe = input_dataframe.drop([0, 1])
    input_dataframe.dropna(axis=0, how='all', inplace=True)
    input_dataframe.reset_index(inplace=True, drop=True)
    input_course_faculty = DataFrame(input_dataframe.iloc[:, : 7])
    input_course_faculty.columns = ['Course Code', 'New Course Title', 'Faculty1', 'Faculty2', 'Faculty3', 'Faculty4',
                                    'Faculty5']
    input_timetable = DataFrame(input_dataframe.iloc[:, 7:])
    input_course_title = input_course_faculty['New Course Title']
    input_course_code = input_course_faculty['Course Code']
    input_timetable.reset_index(inplace=True, drop=True)
    input_course_faculty.reset_index(inplace=True, drop=True)

    # read the latest master calendar copy downloaded from the drive
    existing_master_from_drive = download_master_calendar()
    existing_master_dataframe = pd.read_excel(existing_master_from_drive, sheet_name=month)
    existing_master_dataframe = existing_master_dataframe.drop([0, 1])
    existing_master_dataframe.dropna(axis=0, how='all', inplace=True)
    existing_master_dataframe.reset_index(inplace=True, drop=True)
    existing_master_timetable = DataFrame(existing_master_dataframe.iloc[:, 7:])
    existing_master_course_faculty = DataFrame(existing_master_dataframe.iloc[:, : 7])
    existing_master_course_faculty.columns = ['Course Code', 'New Course Title', 'Faculty1', 'Faculty2', 'Faculty3',
                                              'Faculty4',
                                              'Faculty5']
    existing_course_title = existing_master_course_faculty['New Course Title']
    existing_course_code = existing_master_course_faculty['Course Code']

    # check format whether two slots(new_template = 0) or four slots(new_template = 1)
    if len(input_timetable.columns) > 110:
        new_template = 1
    else:
        new_template = 0

    # reading keys from keysheet
    keys_dataframe = pd.read_excel(input_path, sheet_name='Key')
    keys_dataframe = keys_dataframe.iloc[:, :2]
    keys_dataframe.dropna(inplace=True)
    keys_dataframe.columns = ["title", "code"]
    initiative_title = (keys_dataframe["title"])
    initiative_code = (keys_dataframe["code"])
    initiative_key = {}

    # key dictionary storing title as key and value as code
    for i in range(len(initiative_title)):
        initiative_key[initiative_title[i]] = initiative_code[i]
    set_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    input_timetable.reset_index(inplace=True, drop=True)
    input_course_faculty.reset_index(inplace=True, drop=True)
    input_course_code.reset_index(inplace=True, drop=True)
    input_course_title.reset_index(inplace=True, drop=True)

    # load downloaded master calendar
    WriteExcel = load_workbook(existing_master_from_drive)
    MasterCalendarOutMonth = WriteExcel[month]
    MasterCalendarOutMonth.delete_rows(4, MasterCalendarOutMonth.max_row)

    # check whether there is existing data in master
    is_empty = existing_master_timetable.empty
    if is_empty:
        # total length of output timetable / rows
        total_length = len(input_course_faculty) + len(existing_master_course_faculty)

        # dataframe holding final timetable to be written on master calendar
        output_timetable = pd.DataFrame([[0] * 124] * len(input_course_faculty))

        # if two slots in input, convert to four slots else store in same index
        for i in range(len(input_timetable)):
            for j in range(len(input_timetable.columns)):
                if input_timetable.iloc[i, j] in initiative_key.values():
                    if new_template == 0:
                        output_timetable.iloc[i, j * 2] = input_timetable.iloc[i, j]
                        output_timetable.iloc[i, j * 2 + 1] = input_timetable.iloc[i, j]
                    else:
                        output_timetable.iloc[i, j] = input_timetable.iloc[i, j]
                        output_timetable.iloc[i, j] = input_timetable.iloc[i, j]

        # writing to master first three columns as dataframe and timetable as another dataframe
        rows = dataframe_to_rows(input_course_faculty, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 1):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
        rows = dataframe_to_rows(output_timetable, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 8):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)

    else:
        # holds rows which contain same courses in existing master. Finally these rows are dropped from the dataframe
        # after updating into existing_master_timetable
        drop_data = []

        """
        Check whether input master calendar has same course already present in existing master
        """
        for i in range(0, len(existing_course_code)):  # iterate thru all course code in existing master sheet
            nan_value = 0  # nan_value checks whether input course title empty, if empty, check with course titles
            if pd.isna(
                    existing_course_code[i]) == True:  # if course code empty,check for matching cases with coursetitle
                nan_value = 1
            for j in range(0, len(input_course_code)):  # iterate thru all course code in input master sheet

                if nan_value == 0 and  input_course_title[j] == existing_course_title[i] and input_course_code[j] == existing_course_code[
                    i] :  # check same course code in input and existing master
                    for k in range(
                            len(input_timetable.columns)):  # iterature through all days/timetable for that course
                        # for two slots format and if there is a slot blocked additionally not presenz in existing master
                        if new_template == 0 and input_timetable.iloc[j, k] in initiative_key.values() and \
                                existing_master_timetable.iloc[i, k * 2] != input_timetable.iloc[j, k]:
                            existing_master_timetable.iloc[i, k * 2] = input_timetable.iloc[j, k]
                            existing_master_timetable.iloc[i, (k * 2) + 1] = input_timetable.iloc[j, k]
                        # for four slots format and if there is a slot blocked additionally not present in existing master
                        elif new_template == 1 and (input_timetable.iloc[j, k] in initiative_key.values()) and \
                                existing_master_timetable.iloc[i, k] != input_timetable.iloc[j, k]:
                            existing_master_timetable.iloc[i, k] = input_timetable.iloc[j, k]
                            existing_master_timetable.iloc[i, k] = input_timetable.iloc[j, k]
                    drop_data.append(j)
                # for empty course code, check for matching course title and same as about conditions rest
                elif nan_value == 1 and input_course_title[j] == existing_course_title[i]:
                    existing_course_code[i] = input_course_code[j]
                    for k in range(len(input_timetable.columns)):
                        if new_template == 0 and input_timetable.iloc[j, k] in initiative_key.values() and \
                                existing_master_timetable.iloc[i, k * 2] != input_timetable.iloc[j, k]:
                            existing_master_timetable.iloc[i, k * 2] = input_timetable.iloc[j, k]
                            existing_master_timetable.iloc[i, k * 2 + 1] = input_timetable.iloc[j, k]
                        if new_template == 1 and input_timetable.iloc[j, k] in initiative_key.values() and \
                                existing_master_timetable.iloc[i, k] != input_timetable.iloc[j, k]:
                            existing_master_timetable.iloc[i, k] = input_timetable.iloc[j, k]
                            existing_master_timetable.iloc[i, k] = input_timetable.iloc[j, k]
                    drop_data.append(j)
        drop_data = list(set(drop_data))
        # remove the rows which have same courses in existing master from input dataframe
        for x in drop_data:
            input_timetable = input_timetable.drop([x])
            input_course_faculty = input_course_faculty.drop([x])
        input_timetable.reset_index(inplace=True, drop=True)
        input_course_faculty.reset_index(inplace=True, drop=True)
        total_length = len(input_course_faculty) + len(existing_master_course_faculty)
        output_timetable = pd.DataFrame([[0] * 128] * total_length)

        # update to new timetable format
        for i in range(len(existing_master_timetable)):
            for j in range(len(existing_master_timetable.columns)):
                if existing_master_timetable.iloc[i, j] in initiative_key.values():
                    output_timetable.iloc[i, j] = existing_master_timetable.iloc[i, j]

        # write to excel
        for i in range(len(input_timetable)):
            for j in range(len(input_timetable.columns)):
                if new_template == 0 and input_timetable.iloc[i, j] in initiative_key.values():
                    output_timetable.iloc[len(existing_master_timetable) + i, j * 2] = input_timetable.iloc[i, j]
                    output_timetable.iloc[len(existing_master_timetable) + i, ((j * 2) + 1)] = input_timetable.iloc[
                        i, j]
                elif new_template == 1 and input_timetable.iloc[i, j] in initiative_key.values():
                    output_timetable.iloc[len(existing_master_timetable) + i, j] = input_timetable.iloc[i, j]
                    output_timetable.iloc[len(existing_master_timetable) + i, j] = input_timetable.iloc[i, j]

        output_timetable.drop_duplicates()
        rows = dataframe_to_rows(existing_master_course_faculty, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 1):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
                MasterCalendarOutMonth.border = set_border
        rows = dataframe_to_rows(input_course_faculty, index=False, header=False)
        for r_idx, row in enumerate(rows, 4 + len(existing_master_course_faculty)):
            for c_idx, value in enumerate(row, 1):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
                MasterCalendarOutMonth.border = set_border
        rows = dataframe_to_rows(output_timetable, index=False, header=False)
        for r_idx, row in enumerate(rows, 4):
            for c_idx, value in enumerate(row, 8):
                MasterCalendarOutMonth.cell(row=r_idx, column=c_idx, value=value)
                MasterCalendarOutMonth.border = set_border

    colour_value = ['009EE362', '0000C0D0', '00FFD403', '00FF9356', '007E74D4', '00FE82AA', '00B28DFF', '0085E3FF',
                    '00BFFCC6', '00E7FFAC', '00B5D8D6', '00F6E7E0', '0033CCCC', '00FFCC99', '00333399', '00CC99FF',
                    '00FF00FF', '00FFFF00', '00CCFFFF', '00CCFFCC', '00FFFF99', '0099CCFF']
    key_dictionary = {}
    for i in range(len(initiative_code)):
        key_dictionary[initiative_code[i]] = colour_value[i]

    """Fill colour on session slots"""
    for row in MasterCalendarOutMonth.iter_rows(min_row=4, min_col=8):
        for cell in row:
            cell.border = set_border
            if cell.value in initiative_key.values():
                cell.fill = PatternFill(fill_type='solid',
                                        start_color=key_dictionary[cell.value], end_color=key_dictionary[cell.value])
            if cell.value == 0:
                cell.value = ''

    """ Setting the borders for each cell"""
    for i in range(1, total_length + 4):
        for j in range(1, 124 + 8):
            cell_obj = MasterCalendarOutMonth.cell(row=i, column=j)
            cell_obj.border = set_border
    WriteExcel.save(existing_master_from_drive)
    upload_master_calendar(existing_master_from_drive)




