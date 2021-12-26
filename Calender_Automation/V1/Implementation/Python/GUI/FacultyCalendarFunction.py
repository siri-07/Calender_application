from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import datetime

gauth = GoogleAuth()
gauth.LoadCredentialsFile("credentials.txt")

drive = GoogleDrive(gauth)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd


def upload_file(file_name):
    """Upload file -> change name in setcontentfile"""
    file1 = drive.CreateFile({"mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    file1.SetContentFile(file_name)
    file1.Upload({"convert": True})

    folder_name = 'Download'

    folders = drive.ListFile(
        {
            'q': "title='" + folder_name + "' and mimeType='application/vnd.google-apps.folder' and trashed=false"}).GetList()
    for folder in folders:
        if folder['title'] == folder_name:
            file2 = drive.CreateFile({'parents': [{'id': folder['id']}]})
            file2.SetContentFile(file_name)
            file2.Upload()


def download_master_calendar():
    """
    Searching for a file and downloading it -> insert file name after contains
    give file name to variable file_name
    """
    file_list = drive.ListFile({'q': "title contains 'Master' and trashed=false"}).GetList()
    # print(file_list[0]['title'])
    file_id = file_list[0]['id']
    # print(file_id)
    file = drive.CreateFile({'id': file_id})
    file_time = datetime.now().strftime(" %Y-%m-%d_%I-%M-%S_%p")
    file_name = 'Master'
    file_format = '.xlsx'
    file_title = file_name + file_time + file_format
    file.GetContentFile(file_title,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return file_title


def download_faculty_calendar():
    """
    Searching for a file and downloading it -> insert file name after contains
    give file name to variable file_name
    """
    file_list = drive.ListFile({'q': "title contains 'FacultyCalendar_Output' and trashed=false"}).GetList()
    print(file_list[0]['title'])
    file_id = file_list[0]['id']
    print(file_id)
    file = drive.CreateFile({'id': file_id})
    file_time = datetime.now().strftime(" %Y-%m-%d_%I-%M-%S_%p")
    file_name = 'FacultyCalendar_Output'
    file_format = '.xlsx'
    file_title = file_name + file_time + file_format
    file.GetContentFile(file_title,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return file_title


def FacultyCalendarFunction():
    # Sheet Names
    SheetNames = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October",
                  "November", "December"]
    MasterCalendar = download_master_calendar()  # Download latest Master Calendar
    FacultyCalendar = download_faculty_calendar()  # Download Latest Faculty calendar
    # df = pd.read_excel(MasterCalendar, 'Key')
    # Initiatives = df['FixedInitiativeTitles'].dropna()
    # Initiativecode = df['FixedInitiativeCodes'].dropna()
    # Init = []
    # for i in range(len(Initiatives)):
    # Init.append(Initiativecode[i] + 'p')
    # Init.append(Initiativecode[i] + 's')

    # Running all the sheets of Master Calendar
    for Sheet in range(12):
        MonthSheet = SheetNames[Sheet]
        wb1 = load_workbook(MasterCalendar)
        MasterExcel = wb1[MonthSheet]
        wb2 = load_workbook(FacultyCalendar)
        FacultyExcel = wb2[MonthSheet]
        FacultyExcel.delete_rows(5, FacultyExcel.max_row - 4)
        # Loading all the rows and columns of particular Month sheet
        all_rows = list(MasterExcel.rows)
        all_Columns = list(MasterExcel.columns)
        FacultydList = []
        Leads1 = []
        Leads2 = []
        Leads3 = []
        Leads4 = []
        Leads5 = []
        for row in range(3, len(all_rows)):
            for column in range(7):
                if column == 2:
                    if all_rows[row][column].value:
                        # Getting the Primary Faculty Names
                        FacultydList.append((all_rows[row][column].value).upper())
                        Leads1.append((all_rows[row][column].value).upper())
                    else:
                        Leads1.append((all_rows[row][column].value))
                if column == 3:
                    # Getting the Secondary Faculty Names
                    if all_rows[row][column].value:
                        FacultydList.append((all_rows[row][column].value).upper())
                        Leads2.append((all_rows[row][column].value).upper())
                    else:
                        Leads2.append(all_rows[row][column].value)
                if column == 4:
                    # Getting the Secondary Faculty Names
                    if all_rows[row][column].value:
                        FacultydList.append((all_rows[row][column].value).upper())
                        Leads3.append((all_rows[row][column].value).upper())
                    else:
                        Leads3.append(all_rows[row][column].value)
                if column == 5:
                    # Getting the Secondary Faculty Names
                    if all_rows[row][column].value:
                        FacultydList.append((all_rows[row][column].value).upper())
                        Leads4.append((all_rows[row][column].value).upper())
                    else:
                        Leads4.append(all_rows[row][column].value)
                if column == 6:
                    # Getting the Secondary Faculty Names
                    if all_rows[row][column].value:
                        FacultydList.append((all_rows[row][column].value).upper())
                        Leads5.append((all_rows[row][column].value).upper())
                    else:
                        Leads5.append(all_rows[row][column].value)
        FacultyList = []
        for val in FacultydList:
            if val != None:
                FacultyList.append(val)
        FacultyList = list(set(FacultyList))
        FacultySlots = []
        ExistingSlots = []
        for i in range(len(FacultyList)):
            FacultySlots.append([''] * 124)
        for i in range(len(Leads1)):
            ExistingSlots.append([''] * 124)

        for row in range(3, len(Leads1) + 3):
            for column in range(7, 130):
                if all_rows[row][column].value:
                    ExistingSlots[row - 3][column - 7] = all_rows[row][column].value
        for i in range(len(Leads1)):
            for j in range(len(FacultyList)):
                if FacultyList[j] == Leads1[i]:
                    for k in range(124):
                        if ExistingSlots[i][k]:
                            FacultySlots[j][k] = FacultySlots[j][k] + str(ExistingSlots[i][k]) + 'p'
                if FacultyList[j] == Leads2[i]:
                    for k in range(124):
                        if ExistingSlots[i][k]:
                            FacultySlots[j][k] = FacultySlots[j][k] + str(ExistingSlots[i][k]) + 's'
                if FacultyList[j] == Leads3[i]:
                    for k in range(124):
                        if ExistingSlots[i][k]:
                            FacultySlots[j][k] = FacultySlots[j][k] + str(ExistingSlots[i][k]) + 's'
                if FacultyList[j] == Leads4[i]:
                    for k in range(124):
                        if ExistingSlots[i][k]:
                            FacultySlots[j][k] = FacultySlots[j][k] + str(ExistingSlots[i][k]) + 's'
                if FacultyList[j] == Leads5[i]:
                    for k in range(124):
                        if ExistingSlots[i][k]:
                            FacultySlots[j][k] = FacultySlots[j][k] + str(ExistingSlots[i][k]) + 's'
        ColourValues = ["009EE362", "0000C0D0", "00FFD403", "00FF9356", "007E74D4", "00FE82AA", "00B28DFF", "FF0000"]
        for i in range(len(FacultyList)):
            FacultyExcel.cell(row=5 + i, column=1).value = i + 1
            FacultyExcel.cell(row=5 + i, column=2).value = FacultyList[i]
            correct = 1
            for j in range(124):
                FacultyExcel.cell(row=5 + i, column=3 + j).value = FacultySlots[i][j]
                if FacultySlots[i][j] == "GEp" or FacultySlots[i][j] == "GEs":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[0], end_color=ColourValues[0])
                elif FacultySlots[i][j] == "GE/Pp" or FacultySlots[i][j] == "GE/Ps":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[1], end_color=ColourValues[1])
                elif FacultySlots[i][j] == "B/SUp" or FacultySlots[i][j] == "B/SUs":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[2], end_color=ColourValues[2])
                elif FacultySlots[i][j] == "OTp" or FacultySlots[i][j] == "OTs":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[3], end_color=ColourValues[3])
                elif FacultySlots[i][j] == "Sp" or FacultySlots[i][j] == "Ss":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[4], end_color=ColourValues[4])
                elif FacultySlots[i][j] == "Op" or FacultySlots[i][j] == "Os":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[5], end_color=ColourValues[5])
                elif FacultySlots[i][j] == "COMp" or FacultySlots[i][j] == "COMs":
                    correct = 0
                    # FacultyExcel.cell(row = 5+i, column = 3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[6], end_color=ColourValues[6])
                elif FacultySlots[i][j] != '':
                    FacultyExcel.cell(row=5 + i, column=3 + j).fill = PatternFill(fill_type='solid',
                                                                                  start_color=ColourValues[7],
                                                                                  end_color=ColourValues[7])
        wb2.save(FacultyCalendar)
    upload_file(FacultyCalendar)
