import openpyxl
import re
import matplotlib.pyplot as plt
import pandas as pd
#import plotly.express as px
#import plotly 
    
#-----------fetching the 1st sheet from the (faculty_calendar_july_P1) excel file----


def FacultyLoadSheetFunction():
    #DOWNLOAD FACULTY CALENDAR
    path = "FacultyCalendar_Output.xlsx"
    Faculty_Calendar= openpyxl.load_workbook(path)

    #DOWNLOAD FACULTY LOADSHEET
    FacultyLoadSheet = openpyxl.load_workbook("FacultyLoadSheet.xlsx")

    SheetNames = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

    allBatchs = ["GEs","GEp","B/SUs","B/SUp","OTs","OTp","Ss","Sp","Os","Op","COMs","COMp","GE/Ps","GE/Pp"]

    for Sheet in range(12):
        Faculty_Calendar_Month= Faculty_Calendar[SheetNames[Sheet]]


        load_sheet = FacultyLoadSheet[SheetNames[Sheet]]
        load_sheet.delete_rows(5,load_sheet.max_row-4)
        write_row=1
        m_col = Faculty_Calendar_Month.max_column
        for r in range(5,22):
        
            write_row +=1
            write_col=3
            load_sheet.cell(row = write_row, column =1).value= Faculty_Calendar_Month.cell(row = r, column = 2).value
            for a in range(0,14):
                Count=0
                for c in range(3, m_col+1):
                    cell_obj = Faculty_Calendar_Month.cell(row = r, column = c)
                    val=cell_obj.value
                    s=str(val)
                    #             found=re.compile(allBatchs[a]).match(s)
                    found=re.findall(allBatchs[a],s)
                    if found:
                        Count += 1
                    
                    #         print(Count)
                    #         print(allBatchs[a])
                load_sheet.cell(row = write_row, column =write_col).value=Count
                write_col +=1
                FacultyLoadSheet.save("FacultyLoadSheet.xlsx")
            
        #--------------------------------Calculating total load of faculty------------------------------------------------------------------
        for r1 in range(2,19):
            total_load=0
            for c1 in range(3,17):
                cell_obj1 = load_sheet.cell(row = r1, column = c1)
                v=cell_obj1.value
                total_load=total_load+v
                load_sheet.cell(row =r1, column =2).value=total_load
                FacultyLoadSheet.save("FacultyLoadSheet.xlsx")

    #UPLOAD FACULTY LOADSHEET    
    #---------------------------------------------------------------------------------------------------



